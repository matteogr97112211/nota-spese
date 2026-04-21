from __future__ import annotations

import io
import json
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

import pytesseract
from PIL import Image, ImageFilter, ImageOps
from flask import Flask, Response, render_template, request, send_file
from openpyxl import load_workbook

BASE_DIR = Path(__file__).resolve().parent
TEMPLATE_FILE = BASE_DIR / "Nota Spese Modello.xlsx"

APP_TITLE = "Nota Spese Mobile"
SPESA_TYPES = [
    "Vitto/Alloggio",
    "Varie",
    "Spostamenti",
    "Carburante IT",
    "Cambio € -> Valuta",
    "Cambio Valuta -> €",
    "Prelievo Contante",
]
PAYMENT_TYPES = ["CC-Personale", "Contante", "CC-Visa", "CC-Master", "DKV"]

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024


def sanitize_filename(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return value[:80] or "nota_spese"


def parse_date(value: str | None) -> datetime | None:
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def parse_float(value: str | None) -> float | None:
    if value is None or value == "":
        return None
    value = str(value).replace("€", "").replace(" ", "").replace(",", ".")
    try:
        return float(value)
    except ValueError:
        return None


def normalize_flag(value: str | None) -> str | None:
    return "X" if value and value.upper() == "X" else None


def clear_cell(cell) -> None:
    cell.value = None


def set_riepilogo_fields(wb, data: dict[str, Any]) -> None:
    ws = wb["RIEPILOGO_RIMBORSO_SPESE"]
    mapping = {
        "H6": data.get("rif_ebs"),
        "H7": data.get("autore"),
        "C10": parse_date(data.get("data_inizio")),
        "C11": parse_date(data.get("data_fine")),
        "C12": data.get("cliente"),
        "C13": data.get("luogo"),
        "C14": data.get("tipo"),
        "C15": data.get("causale_trasferta"),
        "C17": data.get("centro_costo"),
        "C18": data.get("dip_esecutore"),
        "C19": data.get("altro_personale_1"),
        "C20": data.get("altro_personale_2"),
        "C21": data.get("altro_personale_3"),
        "C22": data.get("targa_auto_ebs"),
        "C24": parse_float(data.get("notti_italia")),
        "C25": parse_float(data.get("notti_estero")),
        "E10": parse_float(data.get("euro_anticipati")) or 0,
        "E13": parse_float(data.get("euro_resi")) or 0,
        "E17": parse_float(data.get("valuta_anticipata")) or 0,
        "E23": parse_float(data.get("valuta_resa")) or 0,
        "E25": parse_float(data.get("tasso_cambio")) or 1,
        "E16": data.get("tipo_valuta") or "---",
        "J11": parse_float(data.get("bs_causale_1_importo")) or 0,
        "J12": parse_float(data.get("bs_causale_2_importo")) or 0,
        "J13": parse_float(data.get("bs_causale_3_importo")) or 0,
        "G11": data.get("bs_causale_1") or "Confort Viaggi  rif: ",
        "G12": data.get("bs_causale_2") or "TAXI EMMEPI, ",
        "G13": data.get("bs_causale_3") or "AVIS",
    }
    for cell_ref, value in mapping.items():
        ws[cell_ref] = value


def fill_spese_sheet(wb, spese: list[dict[str, Any]]) -> None:
    ws = wb["RIMBORSO_SPESE"]
    for row in range(11, 141):
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            clear_cell(ws[f"{col}{row}"])
    limit = 130
    for idx, item in enumerate(spese[:limit], start=11):
        ws[f"B{idx}"] = parse_date(item.get("data"))
        ws[f"C{idx}"] = item.get("tipo") or None
        ws[f"D{idx}"] = item.get("causale") or None
        ws[f"E{idx}"] = parse_float(item.get("importo_euro"))
        ws[f"F{idx}"] = parse_float(item.get("importo_valuta"))
        ws[f"G{idx}"] = item.get("pagamento") or None
        ws[f"H{idx}"] = normalize_flag(item.get("con_fattura"))
        ws[f"I{idx}"] = normalize_flag(item.get("non_giustificata"))
        ws[f"J{idx}"] = normalize_flag(item.get("rappresentanza"))
        if idx <= 34:
            ws[f"K{idx}"] = normalize_flag(item.get("fiera"))


def fill_km_sheet(wb, tragitti: list[dict[str, Any]]) -> None:
    ws = wb["RIMBORSO KM"]
    for row in range(10, 57):
        ws[f"B{row}"] = None
        ws[f"C{row}"] = None
        ws[f"G{row}"] = None
    limit = 47
    for idx, item in enumerate(tragitti[:limit], start=10):
        ws[f"B{idx}"] = parse_date(item.get("data"))
        ws[f"C{idx}"] = item.get("causale") or None
        ws[f"G{idx}"] = parse_float(item.get("km"))


def workbook_from_form(form: dict[str, Any]) -> io.BytesIO:
    wb = load_workbook(TEMPLATE_FILE)
    set_riepilogo_fields(wb, form)
    fill_spese_sheet(wb, form.get("spese", []))
    fill_km_sheet(wb, form.get("tragitti", []))
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def preprocess_image(file_storage) -> Image.Image:
    image = Image.open(file_storage.stream)
    image = ImageOps.exif_transpose(image).convert("L")
    image = ImageOps.autocontrast(image)
    w, h = image.size
    if max(w, h) < 1800:
        image = image.resize((w * 2, h * 2))
    image = image.filter(ImageFilter.MedianFilter(size=3))
    image = image.filter(ImageFilter.SHARPEN)
    image = image.point(lambda p: 255 if p > 165 else 0)
    return image


def extract_amounts(text: str) -> list[float]:
    candidates = []
    for raw in re.findall(r"(?<!\d)(\d{1,4}(?:[.,]\d{3})*[.,]\d{2})(?!\d)", text):
        cleaned = raw.replace(".", "").replace(",", ".")
        try:
            value = float(cleaned)
        except ValueError:
            continue
        if 0.1 <= value <= 10000:
            candidates.append(value)
    return candidates


def extract_date_iso(text: str) -> str | None:
    patterns = [
        r"\b(\d{2})[\/\-.](\d{2})[\/\-.](\d{4})\b",
        r"\b(\d{2})[\/\-.](\d{2})[\/\-.](\d{2})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        d, m, y = match.groups()
        if len(y) == 2:
            y = "20" + y
        try:
            return datetime(int(y), int(m), int(d)).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def choose_merchant(lines: list[str]) -> str:
    banned = (
        "pagamento", "carta", "bancomat", "iva", "p.iva", "ticket", "totale",
        "subtotal", "subtotale", "resto", "euro", "eur", "contante", "documento",
        "scontrino", "ricevuta", "lotto", "matr.", "codice", "transazione"
    )
    for line in lines[:8]:
        low = line.lower()
        if any(word in low for word in banned):
            continue
        digits = sum(c.isdigit() for c in line)
        if digits > max(4, len(line) // 3):
            continue
        if len(line.strip()) < 3:
            continue
        return line[:80]
    return lines[0][:80] if lines else ""


def suggest_type(text: str) -> str:
    low = text.lower()
    if any(k in low for k in ["hotel", "b&b", "booking", "albergo", "residence"]):
        return "Vitto/Alloggio"
    if any(k in low for k in ["ristorante", "restaurant", "bar", "caffe", "caffè", "trattoria", "pizzeria", "breakfast"]):
        return "Vitto/Alloggio"
    if any(k in low for k in ["taxi", "uber", "metro", "treno", "bus", "parcheggio", "pedaggio", "autostrada"]):
        return "Spostamenti"
    if any(k in low for k in ["fuel", "diesel", "benzina", "gasolio", "eni", "q8", "shell", "tamoil", "esso", "ip "]):
        return "Carburante IT"
    return "Varie"


def build_receipt_payload(text: str) -> dict[str, Any]:
    lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines()]
    lines = [line for line in lines if line]
    date_iso = extract_date_iso(text)
    amounts = extract_amounts(text)

    keyword_amount = None
    m = re.search(r"(?:totale|total|importo|amount|da pagare|pagato)[^\d]{0,12}(\d{1,4}(?:[.,]\d{3})*[.,]\d{2})", text.lower(), flags=re.IGNORECASE)
    if m:
        try:
            keyword_amount = float(m.group(1).replace(".", "").replace(",", "."))
        except ValueError:
            keyword_amount = None

    amount_eur = keyword_amount
    if amount_eur is None and amounts:
        amount_eur = max(amounts)

    merchant = choose_merchant(lines)
    expense_description = merchant or "Spesa da verificare"
    confidence = "media"
    if amount_eur and date_iso and merchant:
        confidence = "alta"
    elif amount_eur or date_iso:
        confidence = "media"
    else:
        confidence = "bassa"

    currency = "EUR"
    lowered = text.lower()
    if re.search(r"\b(?:usd|\$)\b", lowered):
        currency = "USD"
    elif re.search(r"\b(?:gbp|£)\b", lowered):
        currency = "GBP"
    elif re.search(r"\b(?:chf)\b", lowered):
        currency = "CHF"

    return {
        "date_iso": date_iso,
        "merchant": merchant,
        "expense_description": expense_description,
        "suggested_type": suggest_type(text),
        "amount_eur": round(amount_eur, 2) if amount_eur is not None and currency == "EUR" else None,
        "amount_original": round(amount_eur, 2) if amount_eur is not None and currency != "EUR" else None,
        "currency": currency,
        "confidence": confidence,
        "ocr_text_preview": "\n".join(lines[:12]),
    }


@app.get("/")
def index() -> str:
    defaults = {
        "ai_ready": True,
        "tasso_cambio": "1",
        "tipo_valuta": "---",
        "bs_causale_1": "Confort Viaggi  rif: ",
        "bs_causale_2": "TAXI EMMEPI, ",
        "bs_causale_3": "AVIS",
        "spese": [{
            "data": "", "tipo": "", "causale": "", "importo_euro": "",
            "importo_valuta": "", "pagamento": "", "con_fattura": "",
            "non_giustificata": "", "rappresentanza": "", "fiera": "",
        }],
        "tragitti": [{"data": "", "causale": "", "km": ""}],
    }
    return render_template(
        "index.html",
        app_title=APP_TITLE,
        spesa_types=SPESA_TYPES,
        payment_types=PAYMENT_TYPES,
        data=defaults,
        generated=False,
    )


@app.post("/scan-receipt")
def scan_receipt():
    if "receipt" not in request.files:
        return {"error": "File mancante."}, 400
    file = request.files["receipt"]
    if not file.filename:
        return {"error": "Nessun file selezionato."}, 400

    try:
        image = preprocess_image(file)
        text = pytesseract.image_to_string(image, lang="eng", config="--psm 6")
        if not text.strip():
            text = pytesseract.image_to_string(image, lang="eng", config="--psm 11")
        return build_receipt_payload(text)
    except Exception as exc:
        return {"error": f"Errore OCR: {exc}"}, 500


@app.post("/generate")
def generate() -> Response:
    payload_raw = request.form.get("payload")
    if not payload_raw:
        return Response("Dati mancanti.", status=400)
    try:
        payload = json.loads(payload_raw)
    except json.JSONDecodeError:
        return Response("Formato dati non valido.", status=400)

    workbook = workbook_from_form(payload)
    filename_base = sanitize_filename(
        f"nota_spese_{payload.get('cliente') or 'cliente'}_{payload.get('data_inizio') or datetime.now().date()}"
    )
    return send_file(
        workbook,
        as_attachment=True,
        download_name=f"{filename_base}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/health")
def health():
    return "OK", 200


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(host="0.0.0.0", port=port, debug=debug_mode)
